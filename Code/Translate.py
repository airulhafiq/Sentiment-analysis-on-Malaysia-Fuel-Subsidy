import os
import openpyxl
from googletrans import Translator
import time

# Initialize the translator with the correct attribute
translator = Translator(raise_exception=True)  # Ensure 'raise_exception' is used, not 'raise_Exception'

def translate_malay_to_english(malay_text):
    if malay_text is None:
        return ""
    
    if isinstance(malay_text, str):
        if malay_text.strip() == "":
            return ""
    else:
        return malay_text  # If it's not a string, return the original value (e.g., numbers)

    
    try:
        # If it's a valid string, translate it
        translated = translator.translate(malay_text, src='ms', dest='en')
        time.sleep(3) # Add delay of 3 seconds to avoid API rate limits
        return translated.text
    except Exception as e:
        print(f"Translation failed for '{malay_text}': {e}")
        return malay_text  # Return the original text in case of error


# Directories for input and output files
input_folder = 'BM'  # Folder containing the Excel files with Malay text
translated_folder = 'translated'  # Folder to store the translated files

# Create the 'translated' folder if it doesn't exist
if not os.path.exists(translated_folder):
    os.makedirs(translated_folder)

# Loop through each file in the 'BM' directory
for file_name in os.listdir(input_folder):
    if file_name.endswith('.xlsx'):  # Process only .xlsx files
        input_file_path = os.path.join(input_folder, file_name)
        
        # Extract the base name and create the new translated file name
        base_name = os.path.splitext(file_name)[0]
        translated_file_name = f"{base_name}_translated.xlsx"
        output_file_path = os.path.join(translated_folder, translated_file_name)

        # Load the workbook and select the first sheet
        wb = openpyxl.load_workbook(input_file_path)
        sheet = wb.active

        # Create a new workbook to store translated results
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Translated Words"

        # Loop through each row, starting from row 2 (to skip headers if present)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # Ensure the row has at least 2 columns (for column A and B)
            if len(row) > 1:
                original_name = row[0].value  # Copy column A (index 0, Name)
                malay_text = row[1].value     # Access column B (index 1, Malay Text)
                
                new_sheet[f"A{row_num}"] = original_name  # Write the name to column A in the new file
                
                if malay_text is not None:
                    # Translate the Malay text from column B and write to column B in the new file
                    english_translation = translate_malay_to_english(malay_text)
                    new_sheet[f"B{row_num}"] = english_translation

        # Save the new workbook with translations in the 'translated' folder
        new_wb.save(output_file_path)

        print(f"Translation completed for '{file_name}'. Translated file saved as '{output_file_path}'.")

print(f"All translations completed. Translated files are saved in the '{translated_folder}' folder.")
